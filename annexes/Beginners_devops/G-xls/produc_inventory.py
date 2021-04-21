import openpyxl
from datetime import datetime

inv_file       = openpyxl.load_workbook("inventory.xlsx")
product_list   = inv_file["Sheet1"]

today_date     = datetime.today().strftime("%Y-%m-%d")
today_filename = f"inventory_with_total_value_{today_date}.xlsx"

product_per_supplier = {}
price_per_supplier   = {}
product_under_10     = {}
print(product_list.max_row)

def product_per_supplier_calc(product_per_supplier, product_list):
    list_colomn_supplier = range(2, product_list.max_row + 1)
    print(f"DEBUG: product_per_suppllier_calc(), list_colomn_supplier var = {list_colomn_supplier}")
    for product_row in list_colomn_supplier:
        supplier_name = product_list.cell(product_row, 4).value
        inventory     = product_list.cell(product_row, 2).value
        price         = product_list.cell(product_row, 3).value
        product_num   = product_list.cell(product_row, 1).value

        total_price   = product_list.cell(product_row, 5)
        add_supplier_name_filter(product_per_supplier, supplier_name)
        add_price_filter(price_per_supplier, supplier_name, inventory, price)
        product_under_10_filter(product_under_10, product_num, inventory)
        total_price_calc(total_price, inventory, price)

def add_supplier_name_filter(product_per_supplier, supplier_name):
    #print(f"DEBUG: add_supplier_name_filter() supplier_name vaut : {supplier_name}")
    if supplier_name in product_per_supplier:
        current_value = product_per_supplier.get(supplier_name)
        #print(f"DEBUG: add_suppllier_name_filter() current_value vaut {current_value}")
        product_per_supplier[supplier_name] = current_value + 1
    else:
        print("adding a new supplier")
        product_per_supplier[supplier_name] = 1

def add_price_filter(price_per_supplier, supplier_name, inventory, price):
    if supplier_name in price_per_supplier:
        last_total_price = price_per_supplier.get(supplier_name)
        price_per_supplier[supplier_name] = last_total_price + (inventory * price)
    else:
        print(inventory * price)
        price_per_supplier[supplier_name] = inventory * price

def product_under_10_filter(product_under_10, product_num, inventory):
    if inventory < 10:
        product_under_10[int(product_num)] = int(inventory)

def total_price_calc(total_price, inventory, price):
    total_price.value = inventory * price

if __name__ == "__main__":
    product_per_supplier_calc(product_per_supplier, product_list)
    print(product_per_supplier)
    print(price_per_supplier)
    print(product_under_10)

    inv_file.save(today_filename)
